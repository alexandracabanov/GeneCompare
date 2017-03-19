#!/usr/local/bin/python3
import openpyxl

#ws, colname, colsize, bgcolor, bold
def headerize (wsname, colname,colpos,colsize):
 wsname.cell(row=1, column=colpos, value=colname)
 wsname.cell(row=1, column=colpos).fill = openpyxl.styles.PatternFill('solid', 'B9A9A9')
 wsname.cell(row=1, column=colpos).font = openpyxl.styles.Font(bold=True)

 alfa = "abcdefghijklmnopqrstuvwxyz"
 colletter=alfa[colpos-1] 
 wsname.column_dimensions[colletter.upper()].width = colsize


 

