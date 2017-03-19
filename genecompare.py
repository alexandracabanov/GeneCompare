#!/usr/local/bin/python3

import csv
import math
from openpyxl import Workbook
import openpyxl.styles 
from getcolors import getcolor
import argparse
from headerizer import headerize
#openpyxl is a library to manipulate excel files - no real need for two imports just following suit with some sample codes
#getcolors was made by me before I discovered tidyverse
#argparse helps you write command line interface by helping how you handle command line arguments
#headerizer is a function to set formatting of workbook headers - it save a couple lines of code 

#Deal with command line args
parser = argparse.ArgumentParser(description='message files must be tab delimited "MTA-1.0 output". Everything explodes if not')
parser.add_argument('-w','--wholetumor', help='Specify Whole Tumor File',required=True)
parser.add_argument('-c','--cellsonly',help='Specify Tumor Cells Only File', required=True)
parser.add_argument('-o','--output',help='Specifcy the output file name. ".xlsx" will be appended to all names given', required=True)
args = parser.parse_args()


#Open an excel workbook and set WS to the default worksheet
wb = Workbook()
ws1 = wb.active
ws1.title='All-Sorted(logfc var)'
#Set the name of the file that will be written to wbname
wbname=args.output+'.xlsx'
#Create more work sheets and label them
ws2 = wb.create_sheet("LogFC-Sorted(cells only)")
ws3 = wb.create_sheet("FC-Sorted(cells only)")
ws4 = wb.create_sheet("FC-Sorted(fc var)")
ws5 = wb.create_sheet("Hi-Lo)")


#Set the user specific files to file name variables
#Order of the files is critical 
whl_tumor_file=args.wholetumor
tumor_cell_only_file=args.cellsonly

#whl_tumor_file='MTA-1_0.norm.gene.matrix.refseq.BPIvsBP.limma.fdr0.01_fc1.5_whole tumor.txt'
#tumor_cell_only_file='MTA-1_0.norm.gene.matrix.refseq.BPIvsBP.limma.txt.flt.rawp0.01_fc1.5_tumor cells only.txt'

#set variables to column names that will be used for loading a dict
#This speficifies the columns I care about so later it can knows where to pull the data from
col1name='Gene'
col2name='logFC'
col3name='FC'


#open file 1 and create 2 dictionaries for for FC and one for logfc.  Gene is the index for both
#as a reminder r is an order to open read only
with open(whl_tumor_file,'r') as f:
    dict=csv.DictReader(f,delimiter='\t')
    whl_tumor_file_logfc_dict={}
    whl_tumor_file_fc_dict={}
    for key, value in enumerate(dict):
     #get gene, logfc, and fc and set them to variables
     gene=value.get(col1name)
     logchange=value.get(col2name)
     fc=value.get(col3name)
     #load the dictionaries
     whl_tumor_file_logfc_dict[gene.lower()]=logchange #set logfc per gene 
     whl_tumor_file_fc_dict[gene.lower()]=fc #set fc per gene
 
#same as above for file 2
with open(tumor_cell_only_file,'r') as f:
    dict2=csv.DictReader(f,delimiter='\t')
    tumor_cell_only_filedict={}
    tumor_cell_only_filedict2={}
    for key, value in enumerate(dict2):
     gene=value.get(col1name)
     logchange=value.get(col2name)
     fc=value.get(col3name)
     tumor_cell_only_filedict[gene.lower()]=logchange
     tumor_cell_only_filedict2[gene.lower()]=fc


#instantiate some objects 
finaldict={}
nalist=[]
super_dict={}
rowcnt=1
whl_tumor_filefccolor=''
tumor_cell_only_filefccolor=''

#set the header row for stdout and the excel file ws1
row_headers=['Gene','logFC_cells_only','logFC_whole_tumor','logFC Varience','FC_cells_only','FC_whole_tumor','FC Varience']
for i in row_headers:
 print(i,", ",sep='',end='')
print("\r")
cnt=1
for i in row_headers:
 headerize(ws1,i,cnt,15)#headerize is a function i built to set and format row headers. we tell it sheet, cell value, column num, and column width
 cnt=cnt+1
#set headers for ws2
row_headers=['Gene','logFC_cells_only','logFC_whole_tumor']
cnt=1
for i in row_headers:
 headerize(ws2,i,cnt,15)
 cnt=cnt+1
#set headers for ws3
row_headers=['Gene','FC_cells_only','FC_whole_tumor']
cnt=1
for i in row_headers:
 headerize(ws3,i,cnt,15)
 cnt=cnt+1
#set headers for ws4
row_headers=['Gene','FC_cells_only','FC_whole_tumor','FC_Varience']
cnt=1
for i in row_headers:
 headerize(ws4,i,cnt,15)
 cnt=cnt+1


#for loop to compare logfc of each matching gene in the two files
#tumorcellonly file is the critical file to match to the other.  It is the subset to match the the superset
for k,v in tumor_cell_only_filedict.items(): #for key(k) and value(v) in the tumor cell only file dictionary
  whl_tumor_v=whl_tumor_file_logfc_dict.get(k)
  if whl_tumor_v==None:
    nalist.append(k) #add gene name to nalist if it is not found in whole tumore file
  else:
    whl_tumor_v=float(whl_tumor_file_logfc_dict.get(k))
    v=float(v)
    rowcnt+=1
    #get the variance between v and whl_tumor_v
    logfc_delta=abs(float(v) - float(whl_tumor_v))
    
    whl_tumor_filefc=float(whl_tumor_file_fc_dict.get(k))
    tumor_cell_only_filefc=float(tumor_cell_only_filedict2.get(k))
    fc_delta=abs(float(tumor_cell_only_filefc) - float(whl_tumor_filefc))
    finaldict[k]=logfc_delta #not used at present. Cant remember why created
    
    #load dictionary "super_dict" with all the data we will use to make spreadsheets below 
    super_dict.setdefault(k,[]).append(v)   
    super_dict[k].append(whl_tumor_v)   
    super_dict[k].append(logfc_delta)   
    super_dict[k].append(tumor_cell_only_filefc)   
    super_dict[k].append(whl_tumor_filefc)   
    super_dict[k].append(fc_delta)   
    #data_tuple = data_tuple+(k,v,whl_tumor_v,logfc_delta,tumor_cell_only_filefc,whl_tumor_filefc)

    #print results to stdout
    print(k,",",v,",",whl_tumor_v,",",logfc_delta,",",tumor_cell_only_filefc,",",whl_tumor_filefc)



    ''' 
    #create worksheets
    #doesnt seem to make sense to functionize.  Code will still be long
    #create ws1
    #so.. just learned about ws.append... this whole approach needs to be fixed at some point
    ws1.cell(row=rowcnt, column=1, value=k)
    ws1.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws1.cell(row=rowcnt, column=1).number_format = '@' #make text format so excel doesn't make gene names dates.  It is handy getting to borrow from what you've written already
    ws1.cell(row=rowcnt, column=2, value=v)
    ws1.cell(row=rowcnt, column=3, value=whl_tumor_v)
    ws1.cell(row=rowcnt, column=4, value=logfc_delta)
    ws1.cell(row=rowcnt, column=5, value=tumor_cell_only_filefc)
    ws1.cell(row=rowcnt, column=6, value=whl_tumor_filefc)
    #set color formatting per column.  uses function from getcolors.py for logic 
    ws1.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(v)) 
    ws1.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(whl_tumor_v)) 
    ws1.cell(row=rowcnt, column=4).fill = openpyxl.styles.PatternFill('solid', getcolor(logfc_delta)) 
    ws1.cell(row=rowcnt, column=5).fill = openpyxl.styles.PatternFill('solid', getcolor(tumor_cell_only_filefc)) 
    ws1.cell(row=rowcnt, column=6).fill = openpyxl.styles.PatternFill('solid', getcolor(whl_tumor_filefc)) 
    
    #create worksheet2
    ws2.cell(row=rowcnt, column=1, value=k)
    ws2.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=rowcnt, column=1).number_format = '@' 
    ws2.cell(row=rowcnt, column=2, value=v)
    ws2.cell(row=rowcnt, column=3, value=whl_tumor_v)
    ws2.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(v)) 
    ws2.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(whl_tumor_v)) 
    
    #create worksheet3
    ws3.cell(row=rowcnt, column=1, value=k)
    ws3.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws3.cell(row=rowcnt, column=1).number_format = '@'
    ws3.cell(row=rowcnt, column=1).number_format = '@'
    ws3.cell(row=rowcnt, column=2, value=tumor_cell_only_filefc)
    ws3.cell(row=rowcnt, column=3, value=whl_tumor_filefc)
    ws3.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(tumor_cell_only_filefc)) 
    ws3.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(whl_tumor_filefc)) 

    #create worksheet3
    ws4.cell(row=rowcnt, column=1, value=k)
    ws4.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws4.cell(row=rowcnt, column=1).number_format = '@'
    ws4.cell(row=rowcnt, column=2, value=tumor_cell_only_filefc)
    ws4.cell(row=rowcnt, column=3, value=whl_tumor_filefc)
    ws4.cell(row=rowcnt, column=4, value=fc_delta)
    ws4.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(tumor_cell_only_filefc)) 
    ws4.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(whl_tumor_filefc)) 
    ws4.cell(row=rowcnt, column=4).fill = openpyxl.styles.PatternFill('solid', getcolor(fc_delta)) 

ws1.auto_filter.ref = "D2:D95"
ws1.auto_filter.add_sort_condition("D2:D95")

#save the excel file
wb.save(wbname)
'''
#print(super_dict['prex2'][1])
#print(super_dict.keys())

#we're going to great spreadsheets sorted by a paticular field
#so we use sorted function to sort dictionary list items.
#we do this multiple times and store results as sorted objects
logfcdelta_sort=sorted(super_dict.items(), key=lambda x: x[1][2])#ws1
logfccellsonly_sort=sorted(super_dict.items(), key=lambda x: x[1][0])#ws2
logfctumronly_sort=sorted(super_dict.items(), key=lambda x: x[1][1])#used only for top/bottom 5
fccellsonly_sort=sorted(super_dict.items(), key=lambda x: x[1][3])#ws3
fcdelta_sort=sorted(super_dict.items(), key=lambda x: x[1][5])#ws4


#creating work sheets from the contents built above 
rowcnt=1
for i in logfcdelta_sort:
    rowcnt+=1
    ws1.cell(row=rowcnt, column=1, value=i[0]) #gene name 
    ws1.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws1.cell(row=rowcnt, column=1).number_format = '@' #make text format so excel doesn't make gene names dates.
    ws1.cell(row=rowcnt, column=2, value=i[1][0])#values that correspond to column
    ws1.cell(row=rowcnt, column=3, value=i[1][1])
    ws1.cell(row=rowcnt, column=4, value=i[1][2])
    ws1.cell(row=rowcnt, column=5, value=i[1][3])
    ws1.cell(row=rowcnt, column=6, value=i[1][4])
    ws1.cell(row=rowcnt, column=7, value=i[1][5])
    #set color formatting per column.  uses function from getcolors.py for logic 
    ws1.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][0])) 
    ws1.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][1])) 
    ws1.cell(row=rowcnt, column=4).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][2])) 
    ws1.cell(row=rowcnt, column=5).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][3])) 
    ws1.cell(row=rowcnt, column=6).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][4])) 
    ws1.cell(row=rowcnt, column=7).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][5])) 
rowcnt=1
for i in logfccellsonly_sort: 
    rowcnt+=1
    ws2.cell(row=rowcnt, column=1, value=i[0])
    ws2.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws2.cell(row=rowcnt, column=1).number_format = '@' 
    ws2.cell(row=rowcnt, column=2, value=i[1][0])
    ws2.cell(row=rowcnt, column=3, value=i[1][1])
    ws2.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][0])) 
    ws2.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][1])) 
rowcnt=1
for i in fccellsonly_sort: 
    rowcnt+=1
    ws3.cell(row=rowcnt, column=1, value=i[0])
    ws3.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws3.cell(row=rowcnt, column=1).number_format = '@' 
    ws3.cell(row=rowcnt, column=2, value=i[1][3])
    ws3.cell(row=rowcnt, column=3, value=i[1][4])
    ws3.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][3])) 
    ws3.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][4])) 
rowcnt=1
for i in fcdelta_sort:
    rowcnt+=1
    ws4.cell(row=rowcnt, column=1, value=i[0])
    ws4.cell(row=rowcnt, column=1).font = openpyxl.styles.Font(bold=True)
    ws4.cell(row=rowcnt, column=1).number_format = '@' 
    ws4.cell(row=rowcnt, column=2, value=i[1][3])
    ws4.cell(row=rowcnt, column=3, value=i[1][4])
    ws4.cell(row=rowcnt, column=4, value=i[1][5])
    ws4.cell(row=rowcnt, column=2).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][3])) 
    ws4.cell(row=rowcnt, column=3).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][4])) 
    ws4.cell(row=rowcnt, column=4).fill = openpyxl.styles.PatternFill('solid', getcolor(i[1][5])) 

#count how many genes and results we have
cnt=0
for i in logfccellsonly_sort:
 cnt+=1 

#set where we should pull from. Start will always be 0-5, but end is dynamic so its last cnt and last cnt value -5
hstart=0
hend=5
lstart=cnt-5
lend=cnt

#here we get the top/bottom five values using the limits just set
cnt=0
rowcnt=1
for i in logfccellsonly_sort:
 if cnt >= hstart and cnt < hend:
  print(i[0],i[1][0])
  ws5.cell(row=rowcnt, column=1, value=i[1][0])
  rowcnt+=1
 elif cnt >= lstart and cnt <= lend:
  print(i[0],i[1][0])
  ws5.cell(row=rowcnt, column=1, value=i[1][0])
  rowcnt+=1
 cnt+=1

cnt=0
rowcnt=1
for i in logfctumronly_sort:
 if cnt >= hstart and cnt < hend:
  print(i[0],i[1][1])
  ws5.cell(row=rowcnt, column=2, value=i[1][1])
  rowcnt+=1
 elif cnt >= lstart and cnt <= lend:
  print(i[0],i[1][1])
  ws5.cell(row=rowcnt, column=2, value=i[1][1])
  rowcnt+=1
 cnt+=1
wb.save(wbname)


# list.append(i[1]





